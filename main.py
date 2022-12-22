import pandas as pd
import math
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from itertools import chain
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class PrepareData:
    def __init__(self, line, salary):
        self.line = line
        self.salary = salary

    @staticmethod
    def sorter(line):
        line1 = line.copy()
        for i in line1.keys():
            if line1[i] == [] or line1[i] == 0:
                del line[i]
        return dict(sorted(line.items(), key=lambda x: x[1], reverse=True)[:10])

    @staticmethod
    def converter(salary):
        salary_data = salary[0]
        salary_currency = salary[1]
        currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }
        for key, value in currency_to_rub.items():
            if key == salary_currency:
                return math.trunc(value * salary_data)


class Dataset:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy = input('Введите название профессии: ')

    @staticmethod
    def check_input(vacancy):
        if vacancy == 'qwertyuiop':
            return 'abobaboababoa'

    @staticmethod
    def prepare_csv(file_name, vacancy):
        pd.set_option('expand_frame_repr', False)

        df = pd.read_csv(file_name)
        df = df.dropna(subset=df.columns.values)
        df = df[['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']]

        df['published_at'] = df['published_at'].apply(lambda x: int(x[0:4]))
        years = df['published_at'].unique()
        cities = df['area_name'].unique()
        df['salary'] = (df[['salary_from', 'salary_to']].mean(axis=1)).apply(lambda x: math.trunc(x))
        if file_name != 'vacancies_by_year.csv':
            df1 = df[['salary', 'salary_currency']].apply(PrepareData.converter, axis=1)
            df['salary'] = df1
        salary_by_years = {year: [] for year in years}
        vacancies_by_years = {year: 0 for year in years}
        vacancy_salary_by_years = {year: [] for year in years}
        vacancy_counts_by_years = {year: 0 for year in years}
        salary_by_cities = {city: [] for city in cities}
        vacancies_by_cities = {city: 0 for city in cities}

        for year in years:
            salary_by_years[year] = int(df[df['published_at'] == year]['salary'].mean())
            vacancies_by_years[year] = len(df[df['published_at'] == year])

        filtered_df = df[df['name'].str.contains(vacancy)]
        for year in years:
            if vacancy != '':
                vacancy_salary_by_years[year] = int(filtered_df[filtered_df['published_at'] == year]['salary'].mean())
                vacancy_counts_by_years[year] = len(filtered_df[filtered_df['published_at'] == year])
            else:
                vacancy_salary_by_years[year] = 0
                vacancy_counts_by_years[year] = 0

        for city in cities:
            if round((len(df[df['area_name'] == city]) / len(df)), 4) < 0.01:
                continue
            vacancies_by_cities[city] = round((len(df[df['area_name'] == city]) / len(df)), 4)
            salary_by_cities[city] = int(df[df['area_name'] == city]['salary'].mean())
        print('Динамика уровня зарплат по годам:', salary_by_years)
        print('Динамика количества вакансий по годам:', vacancies_by_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', vacancy_salary_by_years)
        print('Динамика количества вакансий по годам для выбранной профессии:', vacancy_counts_by_years)
        print('Уровень зарплат по городам (в порядке убывания):', PrepareData.sorter(salary_by_cities))
        print('Доля вакансий по городам (в порядке убывания):', PrepareData.sorter(vacancies_by_cities))
        return [salary_by_years, vacancies_by_years, vacancy_salary_by_years, vacancy_counts_by_years,
                PrepareData.sorter(salary_by_cities), PrepareData.sorter(vacancies_by_cities)]


class Report(Dataset):
    @staticmethod
    def name_cell(letter, number):
        return f'{letter}{number}'

    @staticmethod
    def cell_generator(list_name, dict_name, letter1, letter2):
        c = 2
        for key, value in dict_name.items():
            list_name[Report.name_cell(letter1, c)] = key
            list_name[Report.name_cell(letter2, c)] = value
            c += 1

    @staticmethod
    def cell_generator_small(list_name, dict_name, letter):
        c = 2
        for i in dict_name.values():
            list_name[Report.name_cell(letter, c)] = i
            c += 1

    @staticmethod
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def set_width(list_name):
        for cell in chain.from_iterable(list_name.iter_cols()):
            if cell.value:
                list_name.column_dimensions[cell.column_letter].width = max(
                    list_name.column_dimensions[cell.column_letter].width,
                    len(f"{cell.value}") + 2,
                )

    @staticmethod
    def get_data():
        a = Dataset()
        return [Dataset.prepare_csv(a.file_name, a.vacancy), a.vacancy]


    @staticmethod
    def generate_excel():
        getdata = Report.get_data()
        data = getdata[0]
        vacancy = getdata[1]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(('Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy}', 'Количество вакансий',
                   f'Количество вакансий - {vacancy}'))
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        ws['E1'].font = Font(bold=True)
        ws.title = "Статистика по годам"
        cities_list = wb.create_sheet('Статистика по городам')
        salary_by_years = data[0]
        vacancies_by_years = data[1]
        vacancy_salary_by_years = data[2]
        vacancy_counts_by_years = data[3]
        salary_by_cities = data[4]
        vacancies_by_cities = data[5]
        for key, value in vacancies_by_cities.items():
            vacancies_by_cities[key] = f'{round(value * 100, 2)}%'
        Report.cell_generator(ws, salary_by_years, 'A', 'B')
        Report.cell_generator_small(ws, vacancies_by_years, 'D')
        Report.cell_generator_small(ws, vacancy_salary_by_years, 'C')
        Report.cell_generator_small(ws, vacancy_counts_by_years, 'E')
        Report.set_border(ws, f'A1:E{len(salary_by_years) + 1}')
        Report.set_width(ws)
        ws.column_dimensions['A'].width = 6

        cities_list.append(('Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'))
        cities_list['A1'].font = Font(bold=True)
        cities_list['B1'].font = Font(bold=True)
        cities_list['D1'].font = Font(bold=True)
        cities_list['E1'].font = Font(bold=True)
        Report.cell_generator(cities_list, salary_by_cities, 'A', 'B')
        Report.cell_generator(cities_list, vacancies_by_cities, 'D', 'E')
        Report.set_border(cities_list, 'A1:B11')
        Report.set_border(cities_list, 'D1:E11')
        Report.set_width(cities_list)
        cities_list.column_dimensions['C'].width = 2
        for i in range(2, 12):
            cities_list[f'E{i}'].alignment = Alignment(horizontal='right')
        wb.save('report.xlsx')
        wb.close()
        return [data, vacancy]

    @staticmethod
    def font_size(ax, size):
        for item in ([ax.title, ax.xaxis.label, ax.yaxis.label] +
                     ax.get_xticklabels() + ax.get_yticklabels()):
            item.set_fontsize(size)

    @staticmethod
    def generate_image():
        getdata = Report.generate_excel()
        data = getdata[0]
        vacancy = getdata[1]
        years_list = data[0].keys()
        salary_list = list(data[0].values())
        vacancies_list = list(data[1].values())
        vacancy_salary_by_years_list = list(data[2].values())
        vacancy_counts_by_years_list = list(data[3].values())
        salary_by_cities_cities1 = list(data[4].keys())
        salary_by_cities_cities = []
        for city in salary_by_cities_cities1:
            city = city.replace('-', '-\n')
            city = city.replace(' ', '\n')
            salary_by_cities_cities.append(city)

        salary_by_cities_salaries = list(data[4].values())
        vacancies_by_cities_cities = list(data[5].keys())
        vacancies_by_cities_segment = list(data[5].values())
        vacancies_by_cities_cities.append('Другие')
        segments = []
        sum = 0
        for i in vacancies_by_cities_segment:
            num = float(i.replace('%', ''))
            segments.append(num)
            sum += num
        segments.append(100 - sum)
        matplotlib.rcParams.update({'font.size': 6})
        gridsize = (2, 2)
        fig = plt.figure(figsize=(9, 9))
        ax1 = plt.subplot2grid(gridsize, (0, 0))
        ax2 = plt.subplot2grid(gridsize, (0, 1))
        ax3 = plt.subplot2grid(gridsize, (1, 0))
        ax4 = plt.subplot2grid(gridsize, (1, 1))

        index = np.arange(len(years_list))
        bw = 0.3
        ax1.bar(index, salary_list, bw, label='средняя з/п')
        ax1.bar(index + bw, vacancy_salary_by_years_list, bw, label=f'средняя з/п {vacancy}')
        ax1.set_xticks(index + 0.5 * bw, years_list)
        ax1.set_xticklabels([str(x) for x in years_list], rotation=90)
        ax1.set_title('Уровень зарплат по годам', fontsize=8)
        ax1.legend(fontsize=8)
        ax1.grid(visible=True, axis='y')
        Report.font_size(ax1, 8)


        index1 = np.arange(len(years_list))
        bw = 0.3
        ax2.bar(index1, vacancies_list, bw, label='Количество вакансий')
        ax2.bar(index1 + bw, vacancy_counts_by_years_list, bw, label=f'Количество вакансий {vacancy}')
        ax2.set_xticks(index1 + 0.5 * bw, years_list)
        ax2.set_xticklabels([str(x) for x in years_list], rotation=90, fontsize=8)
        ax2.set_title('Количество вакансий по годам', fontsize=8)
        ax2.legend(fontsize=8)
        ax2.grid(visible=True, axis='y')
        Report.font_size(ax2, 8)

        ax3.barh(salary_by_cities_cities, salary_by_cities_salaries)
        ax3.invert_yaxis()
        ax3.set_title('Уровень зарплат по городам', fontsize=8)
        ax3.grid(visible=True, axis='x')

        ax4.pie(segments, labels=vacancies_by_cities_cities)
        ax4.set_title('Доля вакансий по городам', fontsize=8)

        plt.savefig('graph.png')
        return vacancy

    @staticmethod
    def generate_pdf():
        vacancy = Report.generate_image()
        pd.set_option('expand_frame_repr', False)
        wb = pd.ExcelFile("report.xlsx")
        df = wb.parse("Статистика по годам")
        df1 = wb.parse('Статистика по городам')

        cities_statistics = f''
        for i in range(len(df)):
            cities_statistics += '          <tr>\n'
            for j in df.iloc[i].values:
                cities_statistics += f'             <td>{j}</td>\n'
            cities_statistics += '          </tr>\n'

        cities_salaries_statistics = f''
        for i in range(len(df1)):
            cities_salaries_statistics += '         <tr>\n'
            for j in df1.iloc[i].values[0:2]:
                cities_salaries_statistics += f'            <td>{j}</td>\n'
            cities_salaries_statistics += '         </tr>\n'

        cities_vacancies_statistics = f''

        for i in range(len(df1)):
            cities_vacancies_statistics += '            <tr>\n'
            for j in df1.iloc[i].values[3:5]:
                cities_vacancies_statistics += f'               <td>{j}</td>\n'
            cities_vacancies_statistics += '            </tr>\n'

        environment = Environment(loader=FileSystemLoader("templates/"))
        template = environment.get_template("ctemplate.html")
        pdf_template = template.render(
            vacancy=vacancy,
            cities_statistics=cities_statistics,
            cities_salaries_statistics=cities_salaries_statistics,
            cities_vacancies_statistics=cities_vacancies_statistics
        )

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


Report.generate_pdf()
